import pandas as pd
import configparser as cp
import psycopg2
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter


config = cp.RawConfigParser()
config_files = ['plfs-2022-2023_2.properties','plfs-2022-2023.properties','plfs-2021-2022.properties','plfs-2020-2021.properties','plfs-2019-2020.properties',
                'plfs-2018-2019.properties','plfs-2017-2018.properties']
print(config_files)



#read property file

for file in config_files:
    print('Currently Read File  :- ' + file)
    config.read(file)
    sheet_year = config.get('master_properties', 'sheet_year')
    sheet_year = config.get('master_properties', 'sheet_year')
    
    ip = config.get('database_details', 'database.ip')
    port = config.get('database_details', 'database.port')
    username = config.get('database_details', 'database.username')
    password = config.get('database_details', 'database.password')
    dbname = config.get('database_details', 'database.dbname')
    connection = psycopg2.connect(database=dbname, user=username, password=password, host=ip, port=port)
    cursor = connection.cursor()
    
    sheet_location = config.get('master_properties', 'sheet_path')
    print('reading excel file:',sheet_location)
    
    
    
    #read property plfs.tables.unique.sheets and start the nested loops
    unique_sheet_count = int(config.get('plfs_tables_for_etl', 'plfs.tables.unique.sheets'))
    
    #below loop for unique sheets only
    for i in range(unique_sheet_count):
        print('Starting for i(Unique Sheets Count):',i)
        #run below loop for blocks that are split across sheets for "each" unique sheet
        block_count = int(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.block.count'))
        header_names = config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.header.names').split(',')
        row_seggregation_level = int(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.rows.seggregation.level.count'))
        col_seggregation_level = int(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.cols.seggregation.level.count'))
        row_start = int(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.row.start'))
        row_end = int(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.row.end'))
        col_start = column_index_from_string(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.col.start'))
        col_end = column_index_from_string(config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.col.end'))
    
        row_seggregation_values_list = []
        row_seggregation_criteria_list = []
        col_seggregation_values_list = []
        col_seggregation_criteria_list = []
        for k in range(int(row_seggregation_level)):
            row_seggregation_values_list.append(((config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.rows.seggregation.'+str(k+1)+'.values')).split(',')))
            row_seggregation_criteria_list.append(((config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.rows.seggregation.'+str(k+1)+'.criteria'))))
        for l in range(int(col_seggregation_level)):
            col_seggregation_values_list.append(((config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.cols.seggregation.'+str(l+1)+'.values')).split(',')))
            col_seggregation_criteria_list.append(((config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.cols.seggregation.'+str(l+1)+'.criteria'))))
    
        #print('row_seggregation_criteria_list:',row_seggregation_criteria_list)
        #print('row_seggregation_values_list:',row_seggregation_values_list)
        
        #print('col_seggregation_criteria_list:',col_seggregation_criteria_list)
        #print('col_seggregation_values_list:',col_seggregation_values_list)
    
        
        for j in range(block_count):
            header_values = config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.block.'+str(j+1)+'.sheet.header.values').split(',')
            sheet_name = config.get('plfs_tables_for_etl', 'plfs.tables.table.'+str(i+1)+'.block.'+str(j+1)+'.sheet')
            input_df = pd.read_excel(sheet_location,sheet_name)
            print('Read Excel Sheet:',sheet_name)
            for row in range(row_start-2,row_end-1):
                for col in range(col_start-1,col_end):
                    insert_query_prefix = "insert into plfs_fact("+""
                    insert_query_suffix = " values("+"'"
                    for m in range(len(row_seggregation_criteria_list)):
                        if row_seggregation_values_list[m][(row+2)-row_start] == 'NULL' :
                             continue  
                        insert_query_prefix = insert_query_prefix + row_seggregation_criteria_list[m] +","
                        insert_query_suffix = insert_query_suffix + row_seggregation_values_list[m][(row+2)-row_start] +"','"
                    for n in range(len(col_seggregation_criteria_list)):
                        if col_seggregation_values_list[n][col-col_start+1] == 'NULL' :
                             continue
                        insert_query_prefix = insert_query_prefix + col_seggregation_criteria_list[n] +","
                        insert_query_suffix = insert_query_suffix + col_seggregation_values_list[n][col-col_start+1]+"','"
                        #print('col_seggregation_criteria_list['+str(n)+']:',col_seggregation_criteria_list[n],' col_seggregation_values_list['+str(n)+']['+str(col-col_start+1)+']:',col_seggregation_values_list[n][col-col_start])
                    
                    
                    indicator_val = input_df.iloc[row].values[col]
                   
                    # print('Inserting, indicator_val['+str(row+2)+']['+str(get_column_letter(col+1))+']:',indicator_val)
    
                    for o in range(len(header_names)):
                        insert_query_prefix = insert_query_prefix + header_names[o] +","
                        insert_query_suffix = insert_query_suffix + header_values[o]+"','"
    
                    insert_query_prefix = insert_query_prefix  +"indicator_value,"
                    rounded_number = round(float(indicator_val), 1)
                    insert_query_suffix = insert_query_suffix + str(rounded_number)+"','"
    
                    insert_query_prefix = insert_query_prefix  +"created_by,"
                    insert_query_suffix = insert_query_suffix + 'system' +"','"
                    insert_query_prefix = insert_query_prefix  +"updated_by,"
                    insert_query_suffix = insert_query_suffix + 'system' +"','"
    
                    
                    insert_query_prefix = insert_query_prefix  +"year,"
                    insert_query_suffix = insert_query_suffix + sheet_year +"','"
    
    
                    plfs_fact_code = 'PLFS_'+sheet_year+'_'+sheet_name+'_'+str(row+2)+'_'+str(get_column_letter(col+1))
                    plfs_fact_code = plfs_fact_code.replace(' ', '')
                    insert_query_prefix = insert_query_prefix + "plfs_fact_code)"
                    insert_query_suffix = insert_query_suffix + plfs_fact_code+"')"
    
                    final_query = insert_query_prefix+insert_query_suffix
                    
                    # print('final_query',final_query)
    
                    cursor.execute(final_query)
    
                    connection.commit()

connection.close()
print('Data insert successfully')

